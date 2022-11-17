from openpyxl import Workbook
from openpyxl import load_workbook
import os
import re
from unicodedata import normalize

# fileContainer里包含当前python脚本目录下，所有的excel文件，处理的时候，写个for循环即可
fileContainer = []
x = os.path.dirname(os.path.abspath(__file__))
file_dir = x

# 移除解析完后的空白行
def remove(sheet, row):
    # iterate the row object
    for cellD in row:
        # check the value of each cell in
        # the row, if any of the value is not
        # None return without removing the row
        if cellD.value is not None:
            return
    # get the row number from the first cell
    # and remove the row
    sheet.delete_rows(row[0].row, 1)


for root, dirs, files in os.walk(file_dir):
    if root == file_dir:
        for file in files:
            # if file[-4:] == "xlsx" or file[-3:] == "xls":
            if file[-4:] == "xlsx":
                fileContainer.append(file_dir + "\\" + file)

# 拿一个excel测试
# firstFile = fileContainer[0]
for firstFile in fileContainer:

    filePath = firstFile
    excelName = filePath[filePath.rindex("\\")+1:filePath.index(".")]

    wbOriginal = load_workbook(firstFile)
    wsOriginal = wbOriginal.active

    wb = Workbook()
    ws = wb.active
    rowInter = 1

    if wsOriginal["A1"].value.find(".") != -1:
        ws["A1"] = "1 测试"
        rowInter += 1

    for cell in wsOriginal["A"]:
        cellCursor = "A" + str(rowInter)
        cellValue = normalize("NFKC", cell.value)
        if cellValue.find("(") != -1:
            ws[cellCursor] = cellValue[: cellValue.index("(")]
        else:
            ws[cellCursor] = cell.value
        rowInter += 1

    wb2 = Workbook()
    ws2 = wb2.active

    iniLine = 0
    for cell in ws["A"]:
        iniLine += 1
        # valueIndex = cell.value.index(" ") + 1
        valueIndex = cell.value.index(re.findall('[\u4e00-\u9fff]+', cell.value)[0][0])
        if cell.value.find(".") != -1:
            ws2.cell(iniLine - cell.value.count("."), cell.value.count(".") + 1).value = cell.value[valueIndex:]

        elif cell.value.find(".") == -1:
            ws2.cell(iniLine, 1).value = cell.value[valueIndex:]

    for row in ws2:
        remove(ws2, row)

    # 合并
    for colNum in range(1, ws2.max_column+1):
        for rowNum in range(2, ws2.max_row+1):
            if ws2.cell(rowNum, colNum).value is None and ws2.cell(rowNum-1, colNum) is not None:
                ws2.merge_cells(start_row=rowNum-1, start_column=colNum, end_row=rowNum, end_column=colNum)

    wb2.save(excelName +"_NEW" +".xlsx")
